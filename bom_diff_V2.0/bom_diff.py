import pandas as pd
import os
import sys       
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ==================== Phase 1 & 2: 硬件底层配置 ====================

# 引脚重映射字典 (Fuzzy Mapping)：自动识别各种稀奇古怪的表头
ALIAS_DICT = {
    'Designator': ['位号', 'designator', 'refdes', 'reference', '编号'],
    'DNP': ['贴片状态', 'dnp', 'nc', '是否贴片', 'status', '空贴']
}

def normalize_columns(df):
    """引脚重映射：将不同版本的列名统一标准化"""
    new_cols = {}
    for col in df.columns:
        col_lower = str(col).strip().lower()
        mapped_name = col
        for std_name, aliases in ALIAS_DICT.items():
            if col_lower in aliases:
                mapped_name = std_name
                break
        new_cols[col] = mapped_name
    df.rename(columns=new_cols, inplace=True)
    return df

def clean_value(val):
    """数字滤波器：消除大小写和空格带来的误差 (例如 '10 uF' 和 '10UF' 视为相同)"""
    return str(val).strip().upper()

def is_dnp(val):
    """
    判断一个电平值是否处于 DNP (空贴) 状态
    加入防误判逻辑，防止将 NCP111 等带有 NC 字符的芯片误判
    """
    s = str(val).strip().upper()
    
    # 1. 强匹配：单元格绝对等于这些状态码
    if s in ['DNP', 'NC', '空贴', 'TRUE']:
        return True
        
    # 2. 弱匹配 (内联后缀)：含有特征码，且带有分隔符 (空格, 斜杠, 下划线, 横杠)
    # 例如识别 '1UF DNP', '10K/NC', '0R_空贴'
    for kw in ['DNP', 'NC', '空贴']:
        if f" {kw}" in s or f"/{kw}" in s or f"_{kw}" in s or f"-{kw}" in s:
            return True
            
    # 3. 弱匹配 (前缀)：例如 'DNP 1UF'
    for kw in ['DNP ', 'NC ', '空贴 ']:
        if s.startswith(kw):
            return True
            
    return False
# ===================================================================

def load_bom_file(filepath):
    """多协议自动加载器"""
    ext = os.path.splitext(filepath)[1].lower()
    try:
        if ext == '.csv':
            df = pd.read_csv(filepath, dtype=str, encoding='gbk', encoding_errors='ignore').fillna('')
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(filepath, dtype=str).fillna('')
        else:
            return pd.DataFrame()
        return normalize_columns(df)
    except Exception as e:
        print(f"读取异常: {e}")
        return pd.DataFrame()

# ==================== Phase 3: 核心状态机与业务逻辑 ====================

def process_boms(old_path, new_path, output_filename):
    old_df = load_bom_file(old_path)
    new_df = load_bom_file(new_path)

    if old_df.empty or new_df.empty or 'Designator' not in old_df.columns or 'Designator' not in new_df.columns:
        messagebox.showerror("致命错误", "文件加载失败，或未能识别到'位号/Designator'列！\n请检查文件是否被占用或格式是否正确。")
        return False

    old_df.set_index('Designator', inplace=True)
    new_df.set_index('Designator', inplace=True)

    old_keys = set(old_df.index)
    new_keys = set(new_df.index)

    added_keys = new_keys - old_keys
    removed_keys = old_keys - new_keys
    common_keys = old_keys & new_keys

    excel_data = []

    # 1. 移除元件
    for k in removed_keys:
        row_data = {'变更类型': '[-] 移除', '位号': k}
        for col in old_df.columns:
            if old_df.loc[k, col] != '':
                row_data[col] = str(old_df.loc[k, col])
        excel_data.append(row_data)

    # 2. 新增元件
    for k in added_keys:
        row_data = {'变更类型': '[+] 新增', '位号': k}
        for col in new_df.columns:
            if new_df.loc[k, col] != '':
                row_data[col] = str(new_df.loc[k, col])
        excel_data.append(row_data)

    # 3. 修改元件 (加入滤波器与 DNP 专项拦截)
    for key in common_keys:
        old_row = old_df.loc[key]
        new_row = new_df.loc[key]
        row_data = {'变更类型': '[*] 修改', '位号': key}
        has_diff = False
        
        all_cols = list(set(old_df.columns) | set(new_df.columns))
        
        for col in all_cols:
            in_old = col in old_df.columns
            in_new = col in new_df.columns
            
            if in_old and in_new:
                val_old = old_row[col]
                val_new = new_row[col]
                
                old_c = clean_value(val_old)
                new_c = clean_value(val_new)
                
                # 调用数字滤波器进行电平比对
                if in_old and in_new:
                    val_old = old_row[col]
                    val_new = new_row[col]
                    
                    old_c = clean_value(val_old)
                    new_c = clean_value(val_new)
                    
                    # 数字滤波器比对出存在差异
                    if old_c != new_c:
                        # 调用鉴相器，提取新旧状态
                        was_dnp = is_dnp(val_old)
                        is_now_dnp = is_dnp(val_new)
                        
                        # 状态机：上升沿触发 (原本有值，现在变为空贴)
                        if not was_dnp and is_now_dnp:
                            row_data['变更类型'] = '[!] 警报: 变为空贴'
                            
                        # 状态机：下降沿触发 (原本是空贴，现在恢复贴片)
                        elif was_dnp and not is_now_dnp:
                            row_data['变更类型'] = '[!] 警报: 恢复贴片'
                            
                        # 如果不是 DNP 的状态切换，它就会默认保持 '[*] 修改'
                        row_data[col] = f"[{val_old}] -> [{val_new}]"
                        has_diff = True
            elif in_old and not in_new:
                row_data[col] = "[新版缺失]"
                has_diff = True
            elif not in_old and in_new:
                row_data[col] = "[旧版缺失]"
                has_diff = True
                
        if has_diff: 
            excel_data.append(row_data)

    if not excel_data:
        messagebox.showinfo("核对完成", "两份 BOM 完全一致，没有发现任何差异！")
        return True

    # ==================== Phase 4: 输出渲染驱动 ====================
    
    result_df = pd.DataFrame(excel_data).fillna('')
    cols = result_df.columns.tolist()
    cols.insert(0, cols.pop(cols.index('变更类型')))
    cols.insert(1, cols.pop(cols.index('位号')))
    result_df = result_df[cols]
    
    # 将警报排在最前面，然后是新增、移除、修改
    result_df.sort_values(by=['变更类型', '位号'], inplace=True)
    result_df.to_excel(output_filename, index=False)

    # 调用 openpyxl 进行自动着色 (LED 状态灯)
    wb = load_workbook(output_filename)
    ws = wb.active
    
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        change_type = ws.cell(row=row, column=1).value
        if not change_type: continue
        
        if '移除' in change_type:
            for cell in ws[row]: cell.fill = red_fill
        elif '新增' in change_type:
            for cell in ws[row]: cell.fill = green_fill
        elif '警报' in change_type:
            for cell in ws[row]: 
                cell.fill = orange_fill
                cell.font = Font(bold=True, color="FF0000")
        elif '修改' in change_type:
            # 修改行：只高亮发生了变化的特定单元格
            for cell in ws[row]:
                if cell.value and '->' in str(cell.value):
                    cell.fill = yellow_fill

    wb.save(output_filename)
    messagebox.showinfo("核对完成", f"差异矩阵已生成并渲染配色！\n保存在: {output_filename}")
    return True

# ==================== Phase 5: HMI 人机交互界面 ====================

def run_app():
    root = tk.Tk()
    root.title("BOM 差异智能核对中枢 V2.0")
    root.geometry("500x250")
    
    old_path_var = tk.StringVar()
    new_path_var = tk.StringVar()

    def select_old():
        filepath = filedialog.askopenfilename(title="选择旧版 BOM (基准)", filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.xlsm *.csv")])
        if filepath: old_path_var.set(filepath)

    def select_new():
        filepath = filedialog.askopenfilename(title="选择新版 BOM (待测)", filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.xlsm *.csv")])
        if filepath: new_path_var.set(filepath)

    def execute():
        old_p = old_path_var.get()
        new_p = new_path_var.get()
        if not old_p or not new_p:
            messagebox.showwarning("警告", "必须同时选择旧版和新版 BOM！")
            return
        
        # ================== 核心寻址修复 ==================
        # 探测当前程序是否处于被 PyInstaller 封装的冻结态 (Frozen State)
        if getattr(sys, 'frozen', False):
            # 如果是双击运行的 .exe，直接抓取这个 .exe 文件所在的绝对物理目录
            base_dir = os.path.dirname(sys.executable)
        else:
            # 如果是在 VS Code 里跑的 .py 源码，抓取这个 .py 所在的绝对物理目录
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
        # 在锁定好的根目录下，拼接出最终的输出文件名
        output_name = os.path.join(base_dir, "BOM_差异智能核对矩阵_V2.xlsx")
        # ==================================================
        
        process_boms(old_p, new_p, output_name)

    # 界面布局搭建 (类似放置元器件)
    tk.Label(root, text="BOM 差异智能核对中枢 V2.0", font=("Arial", 14, "bold")).pack(pady=10)
    
    frame1 = tk.Frame(root)
    frame1.pack(fill='x', padx=20, pady=5)
    tk.Button(frame1, text="1. 加载旧版 BOM", width=15, command=select_old).pack(side='left')
    tk.Entry(frame1, textvariable=old_path_var, state='readonly').pack(side='left', fill='x', expand=True, padx=5)

    frame2 = tk.Frame(root)
    frame2.pack(fill='x', padx=20, pady=5)
    tk.Button(frame2, text="2. 加载新版 BOM", width=15, command=select_new).pack(side='left')
    tk.Entry(frame2, textvariable=new_path_var, state='readonly').pack(side='left', fill='x', expand=True, padx=5)

    tk.Button(root, text="⚡ 开始智能核对 ⚡", bg="lightblue", font=("Arial", 12, "bold"), command=execute).pack(pady=20, fill='x', padx=50)

    root.mainloop()

if __name__ == "__main__":
    run_app()