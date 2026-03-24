import pandas as pd
import os
import sys
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ==================== Module 1: 硬件抽象层 (HAL) 与数字滤波器 ====================

# 引脚重映射字典：包含你见过的所有奇葩表头
ALIAS_DICT = {
    'Designator': ['位号', 'designator', 'refdes', 'reference', '编号'],
    'DNP': ['贴片状态', 'dnp', 'nc', '是否贴片', 'status', '空贴'],
    'K3_Code': ['k3 no.', 'k3 no', 'k3编码', '物料代码', '物料编码', 'k3 code', '编码'],
    'Value': ['value', '值', '参数', '规格型号', '容值', '阻值'],
    'Footprint': ['footprint', '封装', 'package']
}

def clean_param_for_match(val):
    """高级特征滤波器：剥离 AD 祖传的无用前缀，提取核心参数"""
    s = str(val).strip().upper()
    # 剥离 AD 常见的封装前缀，例如 'CAPC-0805' -> '0805', 'RESC-0402' -> '0402'
    s = re.sub(r'^(CAPC|CAPAE|RESC|IND|LED|DIOC|SOP|QFN|SOT)-?', '', s)
    return s

def clean_value(val):
    """底层信号清洗：消除大小写和多余空格"""
    return str(val).strip().upper()

def is_dnp(val):
    """施密特触发器：精准识别空贴状态"""
    s = str(val).strip().upper()
    if s in ['DNP', 'NC', '空贴', 'TRUE']: return True
    for kw in ['DNP', 'NC', '空贴']:
        if f" {kw}" in s or f"/{kw}" in s or f"_{kw}" in s or f"-{kw}" in s: return True
    for kw in ['DNP ', 'NC ', '空贴 ']:
        if s.startswith(kw): return True
    return False

def get_base_dir():
    """物理绝对寻址：防止 EXE 打包后找不到路径"""
    if getattr(sys, 'frozen', False): 
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# ==================== Module 2: 智能总线驱动 (带表头雷达) ====================

def normalize_columns(df):
    """引脚别名规范化引擎"""
    new_cols = {}
    for col in df.columns:
        col_clean = str(col).strip().lower().replace('\ufeff', '')
        mapped_name = col
        for std_name, aliases in ALIAS_DICT.items():
            if col_clean in aliases:
                mapped_name = std_name
                break
        new_cols[col] = mapped_name
    df.rename(columns=new_cols, inplace=True)
    return df

def load_bom_file(filepath):
    """多协议智能加载器 (搭载雷达穿透废话行，防死锁)"""
    ext = os.path.splitext(filepath)[1].lower()
    try:
        # 1. 强行以“无表头”模式加载，防止错位
        if ext == '.csv':
            df = pd.read_csv(filepath, dtype=str, encoding='gbk', encoding_errors='ignore', header=None).fillna('')
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(filepath, dtype=str, header=None).fillna('')
        else:
            return pd.DataFrame()
            
        if df.empty: return df

        # 2. 启动智能雷达扫描表头 (向下扫描30行)
        trigger_words = [a for aliases in ALIAS_DICT.values() for a in aliases]
        header_idx = 0
        
        for i in range(min(30, len(df))):
            row_vals = [str(x).strip().lower().replace('\ufeff', '') for x in df.iloc[i].values]
            if any(cell in trigger_words for cell in row_vals):
                header_idx = i
                break
                
        # 3. 截断无效数据，重构矩阵
        df.columns = df.iloc[header_idx].astype(str).tolist()
        df = df[header_idx + 1:].reset_index(drop=True)
        
        return normalize_columns(df)
        
    except PermissionError:
        messagebox.showerror("总线占用", f"致命错误：文件被占用！\n\n【{os.path.basename(filepath)}】 正被 WPS 或 Excel 打开。\n请立即关闭该文件后重试！")
        return pd.DataFrame()
    except Exception as e:
        messagebox.showerror("读取异常", f"底层错误：\n{e}")
        return pd.DataFrame()

# ==================== Module 3: 核心运算引擎 (双核异构) ====================

def process_diff(old_path, new_path, output_filename):
    """[模式 A]：新旧 BOM 差异对比核对"""
    old_df = load_bom_file(old_path)
    new_df = load_bom_file(new_path)

    if old_df.empty or new_df.empty: return
    if 'Designator' not in old_df.columns or 'Designator' not in new_df.columns:
        messagebox.showerror("位号丢失", "未能识别到'位号 (Designator)'列，请检查文件格式。")
        return

    old_df.set_index('Designator', inplace=True)
    new_df.set_index('Designator', inplace=True)

    old_keys, new_keys = set(old_df.index), set(new_df.index)
    excel_data = []

    # 1. 移除与新增计算
    for k in old_keys - new_keys:
        row = {'变更类型': '[-] 移除', '位号': k}
        for col in old_df.columns: row[col] = str(old_df.loc[k, col])
        excel_data.append(row)

    for k in new_keys - old_keys:
        row = {'变更类型': '[+] 新增', '位号': k}
        for col in new_df.columns: row[col] = str(new_df.loc[k, col])
        excel_data.append(row)

    # 2. 修改项与 DNP 拦截计算
    for key in old_keys & new_keys:
        old_row, new_row = old_df.loc[key], new_df.loc[key]
        row_data = {'变更类型': '[*] 修改', '位号': key}
        has_diff = False
        
        for col in list(set(old_df.columns) | set(new_df.columns)):
            in_old, in_new = col in old_df.columns, col in new_df.columns
            if in_old and in_new:
                val_old, val_new = old_row[col], new_row[col]
                if clean_value(val_old) != clean_value(val_new):
                    was_dnp, is_now_dnp = is_dnp(val_old), is_dnp(val_new)
                    if not was_dnp and is_now_dnp: row_data['变更类型'] = '[!] 警报: 变为空贴'
                    elif was_dnp and not is_now_dnp: row_data['变更类型'] = '[!] 警报: 恢复贴片'
                    row_data[col] = f"[{val_old}] -> [{val_new}]"
                    has_diff = True
            elif in_old and not in_new: row_data[col], has_diff = "[新版缺失此列]", True
            elif not in_old and in_new: row_data[col], has_diff = "[旧版缺失此列]", True
                
        if has_diff: excel_data.append(row_data)

    if not excel_data:
        messagebox.showinfo("核对完成", "完美匹配！两份 BOM 没有任何差异。")
        return

    result_df = pd.DataFrame(excel_data).fillna('')
    cols = result_df.columns.tolist()
    cols.insert(0, cols.pop(cols.index('变更类型')))
    cols.insert(1, cols.pop(cols.index('位号')))
    result_df[cols].sort_values(by=['变更类型', '位号']).to_excel(output_filename, index=False)
    render_excel(output_filename, 'diff')

def process_lib_check(ad_path, lib_path, output_filename):
    """[模式 B]：AD BOM 与 公司 K3 库首次比对校验"""
    ad_df = load_bom_file(ad_path)
    lib_df = load_bom_file(lib_path) 

    if ad_df.empty or lib_df.empty: return
    if 'Designator' not in ad_df.columns or 'K3_Code' not in ad_df.columns:
        messagebox.showerror("识别失败", "AD BOM 中必须包含'位号'与'K3编码'相关列！")
        return
    if 'K3_Code' not in lib_df.columns:
        messagebox.showerror("识别失败", "公司 K3 库中未识别到'编码'列，请检查表头。")
        return

    # 1. 构建高速哈希字典
    lib_dict = {}
    for _, row in lib_df.iterrows():
        k3_code = str(row['K3_Code']).strip()
        if k3_code and k3_code != 'nan':
            specs = [str(row[c]).strip() for c in lib_df.columns if c != 'K3_Code' and str(row[c]).strip() not in ['', 'nan']]
            lib_dict[k3_code] = " | ".join(specs)

    # 2. 链路状态指示灯：防御断路风险
    messagebox.showinfo("底层诊断", f"数据总线加载完毕！\n\nAD 待测物料数: {len(ad_df)} 项\nK3 标准库物料数: {len(lib_dict)} 项\n\n注：如果K3库数量极少，请确认导出的库是否包含了所有元件大类！")

    # 3. 三重寻址交叉碰撞匹配
    excel_data = []
    for _, row in ad_df.iterrows():
        k3_code = str(row.get('K3_Code', '')).strip()
        ad_val = str(row.get('Value', '')).strip()
        ad_foot = str(row.get('Footprint', '')).strip()
        
        row_data = {
            '位号 (Designator)': row.get('Designator', ''),
            'AD K3编码': k3_code,
            'AD 原理图参数': f"Val: {ad_val} | Foot: {ad_foot}",
            '校验状态': '',
            '公司 K3 库标准参数': ''
        }

        # 状态机 A：缺件或彻底找不到
        if k3_code in ['', 'nan', 'NONE']:
            row_data['校验状态'] = '[!] 缺少编码'
        elif k3_code not in lib_dict:
            row_data['校验状态'] = '[x] 库中无此物料'
            
        # 状态机 B：编码存在 -> 进入深度交叉验证
        else:
            k3_specs_str = lib_dict[k3_code]
            k3_specs_upper = k3_specs_str.upper() # 将 K3 规格转化为大写进行底噪抹平
            
            # 调用高级滤波器提取 AD 核心参数
            val_core = clean_param_for_match(ad_val)
            foot_core = clean_param_for_match(ad_foot)
            
            # 逻辑比较器 (Comparator)
            # 如果 AD 里没填参数，默认放行；如果填了，必须在 K3 规格书里找到这个特征码
            match_val = (val_core == '') or (val_core in k3_specs_upper)
            match_foot = (foot_core == '') or (foot_core in k3_specs_upper)
            
            # 与门 (AND Gate) 判断
            if match_val and match_foot:
                row_data['校验状态'] = '[√] 完美匹配'
            else:
                # 捕获冲突并明确报出是哪个引脚对不上
                conflicts = []
                if not match_val: conflicts.append(f"容阻值({val_core})对不上")
                if not match_foot: conflicts.append(f"封装({foot_core})对不上")
                row_data['校验状态'] = f'[!] 参数冲突: {", ".join(conflicts)}'
                
            row_data['公司 K3 库标准参数'] = k3_specs_str
            
        excel_data.append(row_data)

    result_df = pd.DataFrame(excel_data).fillna('')
    result_df.sort_values(by=['校验状态', '位号 (Designator)'], inplace=True)
    result_df.to_excel(output_filename, index=False)
    render_excel(output_filename, 'check')

# ==================== Module 4: 渲染层 (LED 色彩驱动) ====================

def render_excel(filename, mode):
    wb = load_workbook(filename)
    ws = wb.active
    
    if mode == 'diff':
        fills = {'[-]': PatternFill("solid", fgColor="FFCCCC"), '[+]': PatternFill("solid", fgColor="CCFFCC"),
                 '[!]': PatternFill("solid", fgColor="FFCC99"), '[*]': PatternFill("solid", fgColor="FFFF99")}
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row=row, column=1).value)
            if '移除' in val: [setattr(c, 'fill', fills['[-]']) for c in ws[row]]
            elif '新增' in val: [setattr(c, 'fill', fills['[+]']) for c in ws[row]]
            elif '警报' in val: 
                for c in ws[row]: c.fill, c.font = fills['[!]'], Font(bold=True, color="FF0000")
            elif '修改' in val:
                for c in ws[row]:
                    if c.value and '->' in str(c.value): c.fill = fills['[*]']

    elif mode == 'check':
        # [x]无物料标红，[!]冲突警告标橙色并加粗，[√]完美匹配标绿
        fills = {
            '[x]': PatternFill("solid", fgColor="FF9999"), 
            '[!]': PatternFill("solid", fgColor="FFCC00"), # 亮橙色警报 
            '[√]': PatternFill("solid", fgColor="CCFFCC")
        }
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row=row, column=4).value)
            for key in fills:
                if key in val:
                    for c in ws[row]: c.fill = fills[key]
                    if key == '[!]':  # 针对参数冲突，额外加粗标红字体
                        for c in ws[row]: c.font = Font(bold=True, color="FF0000")
                    break
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row=row, column=4).value)
            for key in fills:
                if key in val:
                    [setattr(c, 'fill', fills[key]) for c in ws[row]]
                    break

    wb.save(filename)
    messagebox.showinfo("执行完毕", f"报告已生成并渲染配色！\n保存在: {filename}")

# ==================== Module 5: 人机交互界面 (HMI) ====================

def run_app():
    root = tk.Tk()
    root.title("BOM 智能协同中枢 V4.0 (双核版)")
    root.geometry("600x320")
    
    notebook = ttk.Notebook(root)
    notebook.pack(pady=10, expand=True, fill='both')
    
    # --- Tab A: 差异核对 ---
    tab_a = ttk.Frame(notebook)
    notebook.add(tab_a, text=" 🔄 模式 A：版本差异核对 (改板用) ")
    
    v_old, v_new = tk.StringVar(), tk.StringVar()
    tk.Label(tab_a, text="应用场景：硬件电路改版后，快速提取新增、移除、修改与空贴的器件", fg="gray").pack(pady=10)
    
    f1 = tk.Frame(tab_a); f1.pack(fill='x', padx=20, pady=5)
    tk.Button(f1, text="基准旧版 BOM", width=15, command=lambda: v_old.set(filedialog.askopenfilename())).pack(side='left')
    tk.Entry(f1, textvariable=v_old, state='readonly').pack(side='left', fill='x', expand=True, padx=5)
    
    f2 = tk.Frame(tab_a); f2.pack(fill='x', padx=20, pady=5)
    tk.Button(f2, text="待测新版 BOM", width=15, command=lambda: v_new.set(filedialog.askopenfilename())).pack(side='left')
    tk.Entry(f2, textvariable=v_new, state='readonly').pack(side='left', fill='x', expand=True, padx=5)
    
    tk.Button(tab_a, text="⚡ 开始差异核对", bg="lightblue", font=("Arial", 11, "bold"), 
              command=lambda: process_diff(v_old.get(), v_new.get(), os.path.join(get_base_dir(), "BOM_差异对比报告.xlsx"))).pack(pady=20, fill='x', padx=80)

    # --- Tab B: K3 校验 ---
    tab_b = ttk.Frame(notebook)
    notebook.add(tab_b, text=" 🔍 模式 B：K3 标准库校验 (首件发板用) ")

    v_ad, v_lib = tk.StringVar(), tk.StringVar()
    tk.Label(tab_b, text="应用场景：首次发出 PCBA 样板前，确保图纸中所有器件均在 K3 ERP 库中存在", fg="gray").pack(pady=10)
    
    f3 = tk.Frame(tab_b); f3.pack(fill='x', padx=20, pady=5)
    tk.Button(f3, text="AD 导出 BOM", width=15, command=lambda: v_ad.set(filedialog.askopenfilename())).pack(side='left')
    tk.Entry(f3, textvariable=v_ad, state='readonly').pack(side='left', fill='x', expand=True, padx=5)
    
    f4 = tk.Frame(tab_b); f4.pack(fill='x', padx=20, pady=5)
    tk.Button(f4, text="公司 K3 总库", width=15, command=lambda: v_lib.set(filedialog.askopenfilename())).pack(side='left')
    tk.Entry(f4, textvariable=v_lib, state='readonly').pack(side='left', fill='x', expand=True, padx=5)

    tk.Button(tab_b, text="⚡ 开始 K3 校验", bg="lightgreen", font=("Arial", 11, "bold"), 
              command=lambda: process_lib_check(v_ad.get(), v_lib.get(), os.path.join(get_base_dir(), "K3物料校验报告.xlsx"))).pack(pady=20, fill='x', padx=80)

    root.mainloop()

if __name__ == "__main__":
    run_app()