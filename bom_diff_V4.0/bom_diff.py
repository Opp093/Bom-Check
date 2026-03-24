import pandas as pd
import os
import sys
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ==================== Module 1: 硬件抽象层 (HAL) 与数字滤波器 ====================

# 引脚重映射字典：包含你见过的所有奇葩表头
ALIAS_DICT = {
    'Designator': ['位号', 'designator', 'refdes', 'reference', '编号'],
    'DNP': ['贴片状态', 'dnp', 'nc', '是否贴片', 'status', '空贴'],
    'K3_Code': ['k3 no.', 'k3 no', 'k3编码', '物料代码', '物料编码', 'k3 code', '编码'],
    'Value': ['value', '值', '参数', '规格型号', '容值', '阻值'],
    'Footprint': ['footprint', '封装', 'package']
}

def clean_param_for_match(val):
    """高级特征滤波器：剥离 AD 祖传的无用前缀，提取核心参数"""
    s = str(val).strip().upper()
    # 剥离 AD 常见的封装前缀，例如 'CAPC-0805' -> '0805', 'RESC-0402' -> '0402'
    s = re.sub(r'^(CAPC|CAPAE|RESC|IND|LED|DIOC|SOP|QFN|SOT)-?', '', s)
    return s

def get_footprint_core(val):
    """封装带通滤波器：提取核心尺寸码，滤除后缀如 -C, -R (例如 0402-C -> 0402)"""
    s = str(val).strip().upper()
    # 强制捕获标准的 4 位 SMD 封装码
    m = re.search(r'(0201|0402|0603|0805|1206|1210|2010|2512|3216|3225)', s)
    if m: return m.group(1)
    
    # 滤除 AD 祖传的杂乱前缀和后缀
    s = re.sub(r'^(CAPC|CAPAE|RESC|IND|LED|DIOC|SOP|QFN|SOT|SMA|SMB|SMC)-?', '', s)
    s = re.sub(r'-[CRLM]$', '', s)
    return s

def expand_value(val):
    """等效参数发生器：处理 100nF=0.1uF, 1MR=1M 等单位换算。极致防爆版"""
    val = str(val).strip().upper()
    
    # [核心修复 1：空信号旁路] 如果传入的是空值，直接返回空列表，彻底停止裂变！
    if not val: 
        return []

    val = val.replace('Ω', '') 
    val = re.sub(r'([KMG])R$', r'\1', val) # 1MR -> 1M
    
    eqs = set() # 绝对明确的空集合声明，杜绝 {} 语法糖异常
    eqs.add(val)
    
    if re.match(r'^\d+(\.\d+)?R$', val): 
        eqs.add(val.replace('R', ''))
    elif re.match(r'^\d+(\.\d+)?$', val): 
        eqs.add(val + 'R') 
        
    if re.match(r'^\d+(\.\d+)?[KMG]$', val): 
        eqs.add(val + 'Ω')

    cap_match = re.match(r'^([\d\.]+)(P|N|U|M)F?$', val)
    if cap_match:
        num, unit = float(cap_match.group(1)), cap_match.group(2)
        # 抛弃批量 update，全部采用单发 add 指令，极致防弹
        if unit == 'N':  
            eqs.add(f"{num/1000:g}U"); eqs.add(f"{num/1000:g}UF")
            eqs.add(f"{num*1000:g}P"); eqs.add(f"{num*1000:g}PF")
            eqs.add(f"{num:g}N"); eqs.add(f"{num:g}NF")
        elif unit == 'U': 
            eqs.add(f"{num*1000:g}N"); eqs.add(f"{num*1000:g}NF")
            eqs.add(f"{num:g}U"); eqs.add(f"{num:g}UF")
        elif unit == 'P': 
            eqs.add(f"{num/1000:g}N"); eqs.add(f"{num/1000:g}NF")
            eqs.add(f"{num:g}P"); eqs.add(f"{num:g}PF")
            
    # 全局无 F 后缀补充
    for e in list(eqs):
        # [核心修复 2：空字符绝缘层] 再次确保处理的不是空串
        if not e: continue 
        
        if e.endswith('F'): 
            eqs.add(e[:-1])
        elif e[-1] in ['P', 'N', 'U', 'M']: 
            eqs.add(e + 'F')
        
    return list(eqs)

def clean_value(val):
    """底层信号清洗：消除大小写和多余空格"""
    return str(val).strip().upper()

def is_dnp(val):
    """施密特触发器：精准识别空贴状态"""
    s = str(val).strip().upper()
    if s in ['DNP', 'NC', '空贴', 'TRUE']: return True
    for kw in ['DNP', 'NC', '空贴']:
        if f" {kw}" in s or f"/{kw}" in s or f"_{kw}" in s or f"-{kw}" in s: return True
    for kw in ['DNP ', 'NC ', '空贴 ']:
        if s.startswith(kw): return True
    return False

def get_base_dir():
    """物理绝对寻址：防止 EXE 打包后找不到路径"""
    if getattr(sys, 'frozen', False): 
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# ==================== Module 2: 智能总线驱动 (带表头雷达) ====================

def normalize_columns(df):
    """引脚别名规范化引擎"""
    new_cols = {}
    for col in df.columns:
        col_clean = str(col).strip().lower().replace('\ufeff', '')
        mapped_name = col
        for std_name, aliases in ALIAS_DICT.items():
            if col_clean in aliases:
                mapped_name = std_name
                break
        new_cols[col] = mapped_name
    df.rename(columns=new_cols, inplace=True)
    return df

def load_bom_file(filepath):
    """多协议智能加载器 (搭载雷达穿透废话行，防死锁)"""
    ext = os.path.splitext(filepath)[1].lower()
    try:
        # 1. 强行以“无表头”模式加载，防止错位
        if ext == '.csv':
            df = pd.read_csv(filepath, dtype=str, encoding='gbk', encoding_errors='ignore', header=None).fillna('')
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(filepath, dtype=str, header=None).fillna('')
        else:
            return pd.DataFrame()
            
        if df.empty: return df

        # 2. 启动智能雷达扫描表头 (向下扫描30行)
        trigger_words = [a for aliases in ALIAS_DICT.values() for a in aliases]
        header_idx = 0
        
        for i in range(min(30, len(df))):
            row_vals = [str(x).strip().lower().replace('\ufeff', '') for x in df.iloc[i].values]
            if any(cell in trigger_words for cell in row_vals):
                header_idx = i
                break
                
        # 3. 截断无效数据，重构矩阵
        df.columns = df.iloc[header_idx].astype(str).tolist()
        df = df[header_idx + 1:].reset_index(drop=True)
        
        return normalize_columns(df)
        
    except PermissionError:
        messagebox.showerror("总线占用", f"致命错误：文件被占用！\n\n【{os.path.basename(filepath)}】 正被 WPS 或 Excel 打开。\n请立即关闭该文件后重试！")
        return pd.DataFrame()
    except Exception as e:
        messagebox.showerror("读取异常", f"底层错误：\n{e}")
        return pd.DataFrame()

# ==================== Module 3: 核心运算引擎 (双核异构) ====================

def process_diff(old_path, new_path, output_filename):
    """[模式 A]：新旧 BOM 差异对比核对"""
    old_df = load_bom_file(old_path)
    new_df = load_bom_file(new_path)

    if old_df.empty or new_df.empty: return
    if 'Designator' not in old_df.columns or 'Designator' not in new_df.columns:
        messagebox.showerror("位号丢失", "未能识别到'位号 (Designator)'列，请检查文件格式。")
        return

    old_df.set_index('Designator', inplace=True)
    new_df.set_index('Designator', inplace=True)

    old_keys, new_keys = set(old_df.index), set(new_df.index)
    excel_data = []

    # 1. 移除与新增计算
    for k in old_keys - new_keys:
        row = {'变更类型': '[-] 移除', '位号': k}
        for col in old_df.columns: row[col] = str(old_df.loc[k, col])
        excel_data.append(row)

    for k in new_keys - old_keys:
        row = {'变更类型': '[+] 新增', '位号': k}
        for col in new_df.columns: row[col] = str(new_df.loc[k, col])
        excel_data.append(row)

    # 2. 修改项与 DNP 拦截计算
    for key in old_keys & new_keys:
        old_row, new_row = old_df.loc[key], new_df.loc[key]
        row_data = {'变更类型': '[*] 修改', '位号': key}
        has_diff = False
        
        for col in list(set(old_df.columns) | set(new_df.columns)):
            in_old, in_new = col in old_df.columns, col in new_df.columns
            if in_old and in_new:
                val_old, val_new = old_row[col], new_row[col]
                if clean_value(val_old) != clean_value(val_new):
                    was_dnp, is_now_dnp = is_dnp(val_old), is_dnp(val_new)
                    if not was_dnp and is_now_dnp: row_data['变更类型'] = '[!] 警报: 变为空贴'
                    elif was_dnp and not is_now_dnp: row_data['变更类型'] = '[!] 警报: 恢复贴片'
                    row_data[col] = f"[{val_old}] -> [{val_new}]"
                    has_diff = True
            elif in_old and not in_new: row_data[col], has_diff = "[新版缺失此列]", True
            elif not in_old and in_new: row_data[col], has_diff = "[旧版缺失此列]", True
                
        if has_diff: excel_data.append(row_data)

    if not excel_data:
        messagebox.showinfo("核对完成", "完美匹配！两份 BOM 没有任何差异。")
        return

    result_df = pd.DataFrame(excel_data).fillna('')
    cols = result_df.columns.tolist()
    cols.insert(0, cols.pop(cols.index('变更类型')))
    cols.insert(1, cols.pop(cols.index('位号')))
    result_df[cols].sort_values(by=['变更类型', '位号']).to_excel(output_filename, index=False)
    render_excel(output_filename, 'diff')

def process_lib_check(ad_path, lib_path, output_filename):
    """[模式 B]：AD BOM 与 公司 K3 库首次比对校验"""
    ad_df = load_bom_file(ad_path)
    lib_df = load_bom_file(lib_path) 

    if ad_df.empty or lib_df.empty: return
    if 'Designator' not in ad_df.columns or 'K3_Code' not in ad_df.columns:
        messagebox.showerror("识别失败", "AD BOM 中必须包含'位号'与'K3编码'相关列！")
        return
    if 'K3_Code' not in lib_df.columns:
        messagebox.showerror("识别失败", "公司 K3 库中未识别到'编码'列，请检查表头。")
        return

    # 1. 构建高速哈希字典
    lib_dict = {}
    for _, row in lib_df.iterrows():
        k3_code = str(row['K3_Code']).strip()
        if k3_code and k3_code != 'nan':
            specs = [str(row[c]).strip() for c in lib_df.columns if c != 'K3_Code' and str(row[c]).strip() not in ['', 'nan']]
            lib_dict[k3_code] = " | ".join(specs)

    # 2. 链路状态指示灯：防御断路风险
    messagebox.showinfo("底层诊断", f"数据总线加载完毕！\n\nAD 待测物料数: {len(ad_df)} 项\nK3 标准库物料数: {len(lib_dict)} 项\n\n注：如果K3库数量极少，请确认导出的库是否包含了所有元件大类！")

    # 3. 三重寻址交叉碰撞匹配
    excel_data = []
    for _, row in ad_df.iterrows():
        k3_code = str(row.get('K3_Code', '')).strip()
        ad_val = str(row.get('Value', '')).strip()
        ad_foot = str(row.get('Footprint', '')).strip()
        
        row_data = {
            '位号 (Designator)': row.get('Designator', ''),
            'AD K3编码': k3_code,
            'AD 原理图参数': f"Val: {ad_val} | Foot: {ad_foot}",
            '校验状态': '',
            '公司 K3 库标准参数': ''
        }

        # 状态机 A：缺件或彻底找不到
        if k3_code in ['', 'nan', 'NONE']:
            row_data['校验状态'] = '[!] 缺少编码'
        elif k3_code not in lib_dict:
            row_data['校验状态'] = '[x] 库中无此物料'
            
        # 状态机 B：编码存在 -> 进入深度交叉验证
        else:
            k3_specs_str = lib_dict[k3_code]
            k3_specs_upper = k3_specs_str.upper() # 将 K3 规格转化为大写进行底噪抹平
            
            # --- 通道 1：封装滤波器检测 ---
            foot_core = get_footprint_core(ad_foot)
            match_foot = (foot_core == '') or (foot_core in k3_specs_upper)
            
            # --- 通道 2：参数裂变器扫描 (多波段并发) ---
            val_eqs = expand_value(ad_val) 
            # 只要裂变出的参数（如 0.1UF, 100NF）有任意一个命中 K3，即视作匹配！
            match_val = (ad_val == '') or any(eq in k3_specs_upper for eq in val_eqs)
            
            # --- 与门 (AND Gate) 仲裁判断 ---
            if match_val and match_foot:
                row_data['校验状态'] = '[√] 完美匹配'
            else:
                # 捕获冲突并明确报出是哪个引脚对不上
                conflicts = []
                if not match_val: conflicts.append(f"参数({ad_val})冲突")
                if not match_foot: conflicts.append(f"封装({ad_foot})冲突")
                row_data['校验状态'] = f'[!] {", ".join(conflicts)}'
                
            row_data['公司 K3 库标准参数'] = k3_specs_str
            
        # ==================== [核心修复：焊接数据输出引脚] ====================
        excel_data.append(row_data) 
        # ======================================================================

    # 循环彻底结束后，再将收集到的数据转化为表格并排序
    result_df = pd.DataFrame(excel_data).fillna('')
    result_df.sort_values(by=['校验状态', '位号 (Designator)'], inplace=True)

    result_df = pd.DataFrame(excel_data).fillna('')
    result_df.sort_values(by=['校验状态', '位号 (Designator)'], inplace=True)
    result_df.to_excel(output_filename, index=False)
    render_excel(output_filename, 'check')

# ==================== Module 4: 渲染层 (LED 色彩驱动) ====================

# ==================== Module 4: 渲染层 (莫兰迪护眼色彩驱动) ====================

def render_excel(filename, mode):
    wb = load_workbook(filename)
    ws = wb.active
    
    # 调色盘：启用低饱和度护眼色系 (莫兰迪色系)
    soft_red = PatternFill("solid", fgColor="FADBD8")    # 柔和粉红 (移除 / 找不到物料)
    soft_green = PatternFill("solid", fgColor="D5F5E3")  # 柔和薄荷绿 (新增 / 完美匹配)
    soft_yellow = PatternFill("solid", fgColor="FCF3CF") # 柔和奶黄 (数值修改)
    soft_orange = PatternFill("solid", fgColor="FDEBD0") # 柔和浅橙 (警报 / 参数冲突)
    
    # 柔和警报字体：使用砖红色替代刺眼的纯正红色
    alert_font = Font(bold=True, color="C0392B") 

    if mode == 'diff':
        fills = {'[-]': soft_red, '[+]': soft_green, '[!]': soft_orange, '[*]': soft_yellow}
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row=row, column=1).value)
            if '移除' in val: [setattr(c, 'fill', fills['[-]']) for c in ws[row]]
            elif '新增' in val: [setattr(c, 'fill', fills['[+]']) for c in ws[row]]
            elif '警报' in val: 
                for c in ws[row]: c.fill, c.font = fills['[!]'], alert_font
            elif '修改' in val:
                for c in ws[row]:
                    if c.value and '->' in str(c.value): c.fill = fills['[*]']

    elif mode == 'check':
        fills = {'[x]': soft_red, '[!]': soft_orange, '[√]': soft_green}
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row=row, column=4).value)
            for key in fills:
                if key in val:
                    for c in ws[row]: 
                        c.fill = fills[key]
                        if key == '[!]': c.font = alert_font
                    break

    wb.save(filename)
    messagebox.showinfo("执行完毕", f"🎉 报表生成成功！\n已应用护眼色彩渲染，保存在:\n{filename}")

# ==================== Module 5: 人机交互界面 (现代扁平化 UI) ====================

def run_app():
    root = tk.Tk()
    root.title("BOM 智能协同中枢 V4.0")
    root.geometry("680x420") # 稍微拉宽界面，更显大气
    root.configure(bg="#F2F4F4") # 窗口底色改为极浅的冷灰色
    
    # 启用现代扁平化主题引擎
    style = ttk.Style()
    if 'clam' in style.theme_names():
        style.theme_use('clam')
        
    # 定义全局组件样式
    font_main = ("Microsoft YaHei", 10)
    font_bold = ("Microsoft YaHei", 10, "bold")
    
    style.configure("TNotebook", background="#F2F4F4")
    style.configure("TNotebook.Tab", padding=[20, 8], font=font_bold, background="#E5E8E8")
    style.map("TNotebook.Tab", background=[("selected", "#FFFFFF")], foreground=[("selected", "#2980B9")])
    
    style.configure("TFrame", background="#FFFFFF")
    style.configure("TLabelframe", background="#FFFFFF", font=font_bold, foreground="#34495E")
    style.configure("TLabelframe.Label", background="#FFFFFF")
    style.configure("TButton", padding=6, font=font_main)
    style.configure("TEntry", padding=6)
    
    notebook = ttk.Notebook(root)
    notebook.pack(pady=15, padx=20, expand=True, fill='both')
    
    # ---------------- 选项卡 1：差异核对 ----------------
    tab_a = ttk.Frame(notebook)
    notebook.add(tab_a, text=" 🔄 [模式 A] 硬件改版差异核对 ")
    
    tk.Label(tab_a, text="应用场景：版本迭代时，抓取元器件的新增、移除、数值修改及空贴变化", 
             fg="#7F8C8D", bg="#FFFFFF", font=("Microsoft YaHei", 9)).pack(pady=15)
    
    # 使用 LabelFrame 增加模块的包裹感
    frame_a = ttk.LabelFrame(tab_a, text=" 文件加载配置 ")
    frame_a.pack(fill='x', padx=25, pady=5)
    
    v_old, v_new = tk.StringVar(), tk.StringVar()
    
    f1 = ttk.Frame(frame_a); f1.pack(fill='x', padx=15, pady=10)
    ttk.Button(f1, text="选择基准旧版", width=14, command=lambda: v_old.set(filedialog.askopenfilename())).pack(side='left')
    ttk.Entry(f1, textvariable=v_old, state='readonly').pack(side='left', fill='x', expand=True, padx=10)
    
    f2 = ttk.Frame(frame_a); f2.pack(fill='x', padx=15, pady=10)
    ttk.Button(f2, text="选择待测新版", width=14, command=lambda: v_new.set(filedialog.askopenfilename())).pack(side='left')
    ttk.Entry(f2, textvariable=v_new, state='readonly').pack(side='left', fill='x', expand=True, padx=10)
    
    # 运行按钮定制化 (淡蓝色)
    style.configure("RunA.TButton", font=("Microsoft YaHei", 11, "bold"), background="#AED6F1", foreground="#1B4F72")
    ttk.Button(tab_a, text="⚡ 开始执行差异核对", style="RunA.TButton",
              command=lambda: process_diff(v_old.get(), v_new.get(), os.path.join(get_base_dir(), "BOM_差异对比报告.xlsx"))).pack(pady=25, fill='x', padx=100)

    # ---------------- 选项卡 2：K3 主数据校验 ----------------
    tab_b = ttk.Frame(notebook)
    notebook.add(tab_b, text=" 🔍 [模式 B] 首次发板 K3 物料校验 ")

    tk.Label(tab_b, text="应用场景：新六层板打样前，拦截图纸中未绑定 K3 编码或容阻值/封装错配的器件", 
             fg="#7F8C8D", bg="#FFFFFF", font=("Microsoft YaHei", 9)).pack(pady=15)
    
    frame_b = ttk.LabelFrame(tab_b, text=" 文件加载配置 ")
    frame_b.pack(fill='x', padx=25, pady=5)

    v_ad, v_lib = tk.StringVar(), tk.StringVar()
    
    f3 = ttk.Frame(frame_b); f3.pack(fill='x', padx=15, pady=10)
    ttk.Button(f3, text="AD 导出 BOM", width=14, command=lambda: v_ad.set(filedialog.askopenfilename())).pack(side='left')
    ttk.Entry(f3, textvariable=v_ad, state='readonly').pack(side='left', fill='x', expand=True, padx=10)
    
    f4 = ttk.Frame(frame_b); f4.pack(fill='x', padx=15, pady=10)
    ttk.Button(f4, text="公司 K3 总库", width=14, command=lambda: v_lib.set(filedialog.askopenfilename())).pack(side='left')
    ttk.Entry(f4, textvariable=v_lib, state='readonly').pack(side='left', fill='x', expand=True, padx=10)

    # 运行按钮定制化 (淡绿色)
    style.configure("RunB.TButton", font=("Microsoft YaHei", 11, "bold"), background="#A9DFBF", foreground="#186A3B")
    ttk.Button(tab_b, text="⚡ 开始交叉特征比对", style="RunB.TButton",
              command=lambda: process_lib_check(v_ad.get(), v_lib.get(), os.path.join(get_base_dir(), "K3物料校验报告.xlsx"))).pack(pady=25, fill='x', padx=100)

    root.mainloop()

if __name__ == "__main__":
    run_app()