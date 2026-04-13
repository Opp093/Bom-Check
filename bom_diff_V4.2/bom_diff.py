import pandas as pd
import os
import sys
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ==================== Module 1: 硬件抽象层 (HAL) 与数字滤波器 ====================

# 引脚重映射字典：包含你见过的所有奇葩表头
ALIAS_DICT = {
    'Designator': ['位号', 'designator', 'refdes', 'reference', '编号'],
    'Quantity': ['数量', 'quantity', 'qty', 'count', '用量'], 
    'DNP': ['贴片状态', 'dnp', 'nc', '是否贴片', 'status', '空贴'],
    'K3_Code': ['k3 no.', 'k3 no', 'k3编码', '物料代码', '物料编码', 'k3 code', '编码','k3'],
    'Value': ['value', '值', '参数', '规格型号', '容值', '阻值'],
    'Footprint': ['footprint', '封装', 'package']
}

def clean_param_for_match(val):
    """高级特征滤波器：剥离 AD 祖传的无用前缀，提取核心参数"""
    s = str(val).strip().upper()
    # 剥离 AD 常见的封装前缀，例如 'CAPC-0805' -> '0805', 'RESC-0402' -> '0402'
    s = re.sub(r'^(CAPC|CAPAE|RESC|IND|LED|DIOC|SOP|QFN|SOT)-?', '', s)
    return s

def get_footprint_core(val):
    """封装带通滤波器：提取核心尺寸码，滤除后缀如 -C, -R (例如 0402-C -> 0402)"""
    s = str(val).strip().upper()
    # 强制捕获标准的 4 位 SMD 封装码
    m = re.search(r'(0201|0402|0603|0805|1206|1210|2010|2512|3216|3225)', s)
    if m: return m.group(1)
    
    # 滤除 AD 祖传的杂乱前缀和后缀
    s = re.sub(r'^(CAPC|CAPAE|RESC|IND|LED|DIOC|SOP|QFN|SOT|SMA|SMB|SMC)-?', '', s)
    s = re.sub(r'-[CRLM]$', '', s)
    return s

def expand_value(val):
    """等效参数发生器：处理 100nF=0.1uF 等单位换算。极致防爆版"""
    val = str(val).strip().upper()
    
    # [空信号旁路] 如果传入的是空值，直接返回空列表
    if not val: 
        return []

    # ================= [核心修复：精度符号对称清洗] =================
    # 抹除 AD 参数中的 ± 和 +/-，使其与 K3 库实现绝对对称匹配
    val = val.replace('±', '').replace('+/-', '')
    # ================================================================

    val = val.replace('Ω', '') 
    val = re.sub(r'([KMG])R$', r'\1', val) # 1MR -> 1M
    
    eqs = set() # 绝对明确的空集合声明，杜绝 {} 语法糖异常
    eqs.add(val)
    
    if re.match(r'^\d+(\.\d+)?R$', val): 
        eqs.add(val.replace('R', ''))
    elif re.match(r'^\d+(\.\d+)?$', val): 
        eqs.add(val + 'R') 
        
    if re.match(r'^\d+(\.\d+)?[KMG]$', val): 
        eqs.add(val + 'Ω')

    cap_match = re.match(r'^([\d\.]+)(P|N|U|M)F?$', val)
    if cap_match:
        num, unit = float(cap_match.group(1)), cap_match.group(2)
        # 抛弃批量 update，全部采用单发 add 指令，极致防弹
        if unit == 'N':  
            eqs.add(f"{num/1000:g}U"); eqs.add(f"{num/1000:g}UF")
            eqs.add(f"{num*1000:g}P"); eqs.add(f"{num*1000:g}PF")
            eqs.add(f"{num:g}N"); eqs.add(f"{num:g}NF")
        elif unit == 'U': 
            eqs.add(f"{num*1000:g}N"); eqs.add(f"{num*1000:g}NF")
            eqs.add(f"{num:g}U"); eqs.add(f"{num:g}UF")
        elif unit == 'P': 
            eqs.add(f"{num/1000:g}N"); eqs.add(f"{num/1000:g}NF")
            eqs.add(f"{num:g}P"); eqs.add(f"{num:g}PF")
            
    # 全局无 F 后缀补充
    for e in list(eqs):
        if not e: continue 
        if e.endswith('F'): 
            eqs.add(e[:-1])
        elif e[-1] in ['P', 'N', 'U', 'M']: 
            eqs.add(e + 'F')
        
    return list(eqs)

def clean_value(val):
    """底层信号清洗：消除大小写和多余空格"""
    return str(val).strip().upper()

def is_dnp(val):
    """施密特触发器：精准识别空贴状态"""
    s = str(val).strip().upper()
    if s in ['DNP', 'NC', '空贴', 'TRUE']: return True
    for kw in ['DNP', 'NC', '空贴']:
        if f" {kw}" in s or f"/{kw}" in s or f"_{kw}" in s or f"-{kw}" in s: return True
    for kw in ['DNP ', 'NC ', '空贴 ']:
        if s.startswith(kw): return True
    return False

def normalize_k3_specs(specs):
    """K3 库参数清洗：与 AD 参数保持同等符号处理，避免非对称误判"""
    return normalize_resistance_unit_text(specs).replace('±', '').replace('+/-', '')

def strip_dnp_noise_from_value(val):
    """从 AD Value 中剥离 DNP/NC/空贴噪音，保留芯片型号中的 NC 字符"""
    clean_ad_val = str(val).strip().upper()
    for kw in ['DNP', 'NC']:
        clean_ad_val = re.sub(r'(?<![A-Z0-9])' + kw + r'(?![A-Z0-9])', '', clean_ad_val)
    clean_ad_val = re.sub(r'空贴', '', clean_ad_val)
    clean_ad_val = re.sub(r'\(\s*\)', '', clean_ad_val)
    clean_ad_val = re.sub(r'(^|[\s_,\/\|;；，]+)-+', r'\1', clean_ad_val)
    return re.sub(r'^[\s_,\/\|\-]+|[\s_,\/\|\-]+$', '', clean_ad_val)

def split_value_tokens(val):
    """按 AD 常见分隔符拆分 Value，过滤空片段"""
    return [s.strip() for s in re.split(r'[_,\/\|\s]+', strip_dnp_noise_from_value(val)) if s.strip()]

def find_value_conflicts(ad_val, k3_specs_upper):
    """返回 AD Value 中未能在 K3 参数文本里命中的切片"""
    val_conflicts = []
    for sv in split_value_tokens(ad_val):
        sv_eqs = {key for key in build_value_search_keys(sv) if not re.fullmatch(r'\d+(\.\d+)?', key)}
        if re.fullmatch(r'[±+\-]?\d+(\.\d+)?%', sv):
            continue
        if re.fullmatch(r'\d+(\.\d+)?[VAW]', sv):
            continue
        if not sv_eqs:
            text_token = normalize_resistance_unit_text(sv)
            if text_token and text_token not in k3_specs_upper:
                val_conflicts.append(sv)
            continue
        if not any(eq in k3_specs_upper for eq in sv_eqs):
            val_conflicts.append(sv)
    return val_conflicts

def value_matches_k3(ad_val, k3_specs_upper):
    """判断 AD Value 是否被 K3 参数覆盖"""
    return len(find_value_conflicts(ad_val, k3_specs_upper)) == 0

def footprint_matches_k3(ad_foot, k3_specs_upper):
    """判断 AD Footprint 的核心封装码是否被 K3 参数覆盖"""
    foot_core = get_footprint_core(ad_foot)
    return (foot_core == '') or (foot_core in k3_specs_upper)

def normalize_resistance_unit_text(val):
    """把 10Ω / 5.6KΩ 规范成 10R / 5.6K，便于等效匹配"""
    s = str(val).strip().upper()
    s = re.sub(r'(\d+(?:\.\d+)?)([KMG])\s*Ω', r'\1\2', s)
    s = re.sub(r'(\d+(?:\.\d+)?)\s*Ω', r'\1R', s)
    return s

def extract_electrical_value_tokens(val):
    """只提取真正的电气主值，避免把 10%、50V、1/10W 等规格噪声当成 Value"""
    raw = normalize_resistance_unit_text(val)
    raw = raw.replace('+/-', ' ').replace('±', ' ')
    raw = re.sub(r'^\s*\(([^()]*)\)\s*$', r'\1', raw)
    raw = re.sub(r'\([^)]*\)', ' ', raw)
    raw = re.sub(r'^\s*(0201|0402|0603|0805|1206|1210|2010|2512|3216|3225)\s*[-_/ ]\s*', '', raw)

    tokens = []
    value_pattern = re.compile(r'(?<![A-Z0-9.])\d+(?:\.\d+)?(?:R|K|M|G|PF|NF|UF|MF|P|N|U|PH|NH|UH|MH|H)(?=[^A-Z0-9]|$)')
    for token in value_pattern.findall(raw):
        if token not in tokens:
            tokens.append(token)

    if not tokens and re.match(r'^\s*0\s*(?:[-_/\s]|$)', raw):
        tokens.append('0R')

    return tokens

def expand_inductance_value(val):
    """电感等效值转换：2.2uH = 2200nH，仅处理带 H 的电感单位"""
    token = str(val).strip().upper()
    m = re.match(r'^([\d\.]+)(P|N|U|M)?H$', token)
    if not m:
        return [token]

    num = float(m.group(1))
    unit = m.group(2) or ''
    eqs = {token}

    if unit == 'U':
        eqs.add(f"{num:g}UH")
        eqs.add(f"{num*1000:g}NH")
    elif unit == 'N':
        eqs.add(f"{num:g}NH")
        eqs.add(f"{num/1000:g}UH")
    elif unit == 'M':
        eqs.add(f"{num:g}MH")
        eqs.add(f"{num*1000:g}UH")
    elif unit == '':
        eqs.add(f"{num:g}H")
        eqs.add(f"{num*1000000:g}UH")

    return list(eqs)

def expand_electrical_value_token(token):
    """按单位类型生成等效主值，避免把 uH 当成 uF"""
    token = str(token).strip().upper()
    if token.endswith('H'):
        return expand_inductance_value(token)
    return expand_value(token)

def extract_tolerance_tokens(val):
    """提取精度，例如 1%、5%、±10%，用于电阻等需要精度唯一性的类别"""
    text = str(val).strip().upper().replace('±', '')
    return sorted(set(re.findall(r'\d+(?:\.\d+)?%', text)))

def get_required_tolerances(ad_val, family):
    """只有明确要求精度唯一性的类别才强制校验精度"""
    if family != 'resistor':
        return []
    return extract_tolerance_tokens(ad_val)

def get_tolerance_conflicts(ad_val, family, k3_specs_upper):
    """返回 K3 中缺失的 AD 精度要求"""
    return [tol for tol in get_required_tolerances(ad_val, family) if tol not in k3_specs_upper]

def build_value_search_keys(val):
    """为主参数建立可查询的等效值集合，例如 100nF 同时索引 0.1uF"""
    keys = set()
    for token in extract_electrical_value_tokens(val):
        keys.add(token)
        keys.update(expand_electrical_value_token(token))
    return {k for k in keys if k}

def add_candidate_index(index, value_key, foot_core, k3_code):
    """写入候选索引，使用集合避免大库构建时反复线性去重"""
    if not value_key:
        return
    index.setdefault((value_key, foot_core), set()).add(k3_code)

def find_value_footprint_k3_codes(ad_val, ad_foot, designator_text, candidate_index, value_index, code_rank, code_category, code_specs):
    """用 AD Value + Footprint 通过索引反查所有适配的 K3 编码"""
    ad_key_sets = []
    for token in split_value_tokens(ad_val):
        keys = build_value_search_keys(token)
        if keys:
            ad_key_sets.append(keys)

    if not ad_key_sets:
        return []

    foot_core = get_footprint_core(ad_foot)
    family = get_designator_family(designator_text)
    token_code_sets = []

    for keys in ad_key_sets:
        token_codes = set()
        for key in keys:
            if foot_core and family in ['resistor', 'capacitor']:
                token_codes.update(candidate_index.get((key, foot_core), []))
            else:
                token_codes.update(value_index.get((key, ''), []))

        if not token_codes:
            return []
        token_code_sets.append(token_codes)

    matched_codes = set.intersection(*token_code_sets)
    sorted_codes = sorted(matched_codes, key=lambda code: code_rank.get(code, len(code_rank)))
    sorted_codes = filter_candidate_codes_by_designator(sorted_codes, designator_text, code_category)

    required_tolerances = get_required_tolerances(ad_val, family)
    if required_tolerances:
        sorted_codes = [
            code for code in sorted_codes
            if all(tol in normalize_k3_specs(code_specs.get(code, '')) for tol in required_tolerances)
        ]

    return sorted_codes

def normalize_k3_code_for_lookup(code):
    """仅用于查库的 K3 编码归一化，兼容 025-GJ / 025GJ 这类连字符差异"""
    return re.sub(r'[\s\-]+', '', str(code).strip().upper())

def split_k3_codes(code_text):
    """拆分 AD 中可能合并在一个单元格里的多个 K3 编码"""
    return [c.strip() for c in re.split(r'[,，;；]+', str(code_text)) if c.strip()]

def resolve_k3_codes(code_text, lib_dict, lib_code_lookup):
    """按原始编码优先、归一化编码兜底，解析 AD K3 到 K3 库真实编码"""
    resolved_codes = []
    missing_codes = []

    for code in split_k3_codes(code_text):
        lib_code = code if code in lib_dict else lib_code_lookup.get(normalize_k3_code_for_lookup(code))
        if lib_code:
            if lib_code not in resolved_codes:
                resolved_codes.append(lib_code)
        else:
            missing_codes.append(code)

    return resolved_codes, missing_codes

def get_designator_prefixes(designator_text):
    """提取位号前缀，例如 R77 -> R、FB1 -> FB"""
    prefixes = set()
    for item in re.split(r'[,，\s]+', str(designator_text).upper()):
        m = re.match(r'([A-Z]+)\d+', item.strip())
        if m:
            prefixes.add(m.group(1))
    return prefixes

def get_designator_family(designator_text):
    """按位号判断物料大类，用于过滤第二轮候选 K3"""
    prefix_family = {
        'R': 'resistor', 'RN': 'resistor',
        'C': 'capacitor', 'CN': 'capacitor',
        'L': 'inductor',
        'FB': 'ferrite',
    }
    families = {prefix_family[p] for p in get_designator_prefixes(designator_text) if p in prefix_family}
    return next(iter(families)) if len(families) == 1 else ''

def get_lib_category(row, spec_text):
    """从 K3 名称/规格中提取物料大类，辅助候选过滤"""
    name_text = str(row.get('名称', '')).strip()
    text = f"{name_text} | {spec_text}"
    if '磁珠' in text:
        return 'ferrite'
    if '电阻' in text:
        return 'resistor'
    if '电容' in text:
        return 'capacitor'
    if '电感' in text:
        return 'inductor'
    return ''

def filter_candidate_codes_by_designator(codes, designator_text, code_category):
    """根据位号类型过滤候选，避免 R 位号里混入磁珠/电容等 K3"""
    family = get_designator_family(designator_text)
    if not family:
        return codes
    return [code for code in codes if code_category.get(code, '') == family]

def check_k3_specs_match(ad_val, ad_foot, designator_text, k3_specs_upper):
    """单个 K3 编码与 AD Value/Footprint 的独立匹配结果"""
    designator_upper = str(designator_text).strip().upper()
    foot_core = get_footprint_core(ad_foot)
    is_rc_passive = re.match(r'^(R|C|RN|CN)\d+', designator_upper)

    if is_rc_passive:
        match_foot = (foot_core == '') or (foot_core in k3_specs_upper)
    else:
        match_foot = True

    val_conflicts = find_value_conflicts(ad_val, k3_specs_upper)
    tolerance_conflicts = get_tolerance_conflicts(ad_val, get_designator_family(designator_text), k3_specs_upper)
    match_val = len(val_conflicts) == 0 and len(tolerance_conflicts) == 0
    return match_val, match_foot, val_conflicts, tolerance_conflicts

def summarize_k3_mismatches(k3_results, ad_foot):
    """生成多 K3 独立校验后的简短冲突摘要"""
    conflicts = []
    for code, match_val, match_foot, val_conflicts, tolerance_conflicts in k3_results:
        if match_val and match_foot:
            continue
        parts = []
        if not match_val:
            detail_parts = []
            if val_conflicts:
                detail_parts.append(','.join(val_conflicts))
            if tolerance_conflicts:
                detail_parts.append("精度" + ','.join(tolerance_conflicts))
            parts.append(f"参数({';'.join(detail_parts)})")
        if not match_foot:
            parts.append(f"封装({ad_foot})")
        conflicts.append(f"{code}:{'/'.join(parts)}")
    return "；".join(conflicts)

def get_base_dir():
    """物理绝对寻址：防止 EXE 打包后找不到路径"""
    if getattr(sys, 'frozen', False): 
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# ==================== Module 2: 智能总线驱动 (带表头雷达) ====================

def normalize_columns(df):
    """引脚别名规范化引擎"""
    new_cols = {}
    for col in df.columns:
        col_clean = str(col).strip().lower().replace('\ufeff', '')
        mapped_name = col
        for std_name, aliases in ALIAS_DICT.items():
            if col_clean in aliases:
                mapped_name = std_name
                break
        new_cols[col] = mapped_name
    df.rename(columns=new_cols, inplace=True)
    return df

def load_bom_file(filepath):
    """多协议智能加载器 (搭载雷达穿透废话行，防死锁)"""
    ext = os.path.splitext(filepath)[1].lower()
    try:
        # 1. 强行以“无表头”模式加载，防止错位
        if ext == '.csv':
            df = pd.read_csv(filepath, dtype=str, encoding='gbk', encoding_errors='ignore', header=None).fillna('')
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(filepath, dtype=str, header=None).fillna('')
        else:
            return pd.DataFrame()
            
        if df.empty: return df

        # 2. 启动智能雷达扫描表头 (向下扫描30行)
        trigger_words = [a for aliases in ALIAS_DICT.values() for a in aliases]
        header_idx = 0
        
        for i in range(min(30, len(df))):
            row_vals = [str(x).strip().lower().replace('\ufeff', '') for x in df.iloc[i].values]
            if any(cell in trigger_words for cell in row_vals):
                header_idx = i
                break
                
        # 3. 截断无效数据，重构矩阵
        df.columns = df.iloc[header_idx].astype(str).tolist()
        df = df[header_idx + 1:].reset_index(drop=True)
        
        return normalize_columns(df)
        
    except PermissionError:
        messagebox.showerror("总线占用", f"致命错误：文件被占用！\n\n【{os.path.basename(filepath)}】 正被 WPS 或 Excel 打开。\n请立即关闭该文件后重试！")
        return pd.DataFrame()
    except Exception as e:
        messagebox.showerror("读取异常", f"底层错误：\n{e}")
        return pd.DataFrame()

# ==================== [防线升级：深度拆包与矩阵摊平引擎] ====================
def flatten_bom_matrix(df):
    """
    底层算法：将压缩的位号字符串 (如 R1, R2, R3) 瞬间爆炸拆解为独立的器件行
    """
    if df.empty or 'Designator' not in df.columns:
        return df
        
    df = df.copy()
    
    # 1. 切割位号：兼容中英文逗号、空格，并过滤掉空值和 pandas 产生的 NAN
    df['Designator'] = df['Designator'].apply(
        lambda x: [d.strip().upper() for d in re.split(r'[,，\s]+', str(x)) 
                   if d.strip() and d.strip().upper() not in ['NAN', 'NONE', 'NULL']]
    )
    
    # 2. 核心大招：矩阵爆炸 (Explode)！把 1 行压缩包变成 100 行物理清单
    df = df.explode('Designator').reset_index(drop=True)
    
    # 3. 拦截空位号
    df = df[df['Designator'].notna()]
    
    # 4. [物理防呆] 强制销毁 Quantity 数量列
    # 防止原本 100 变成 99 时，剩下的 99 个器件因为这 1 个数字的变化被集体误判为“修改”
    if 'Quantity' in df.columns:
        df.drop(columns=['Quantity'], inplace=True)
        
    # 5. 绝对防撞墙：如果工程师在画板时手误留下了重复的位号，强行去重保命
    df.drop_duplicates(subset=['Designator'], keep='first', inplace=True)
    
    return df
# ============================================================================

# ==================== Module 3: 核心运算引擎 (双核异构) ====================

def process_diff(old_path, new_path, output_filename):
    """[模式 A]：新旧 BOM 差异对比核对"""
    old_df = load_bom_file(old_path)
    new_df = load_bom_file(new_path)

    if old_df.empty or new_df.empty: return
    if 'Designator' not in old_df.columns or 'Designator' not in new_df.columns:
        messagebox.showerror("位号丢失", "未能识别到'位号 (Designator)'列，请检查文件格式。")
        return

    # ================= [激活防线：注入摊平引擎] =================
    old_df = flatten_bom_matrix(old_df)
    new_df = flatten_bom_matrix(new_df)
    # ============================================================

    old_df.set_index('Designator', inplace=True)
    new_df.set_index('Designator', inplace=True)

    old_keys, new_keys = set(old_df.index), set(new_df.index)
    excel_data = []
    

    # 1. 移除与新增计算
    for k in old_keys - new_keys:
        row = {'变更类型': '[-] 移除', '位号': k}
        for col in old_df.columns: row[col] = str(old_df.loc[k, col])
        excel_data.append(row)

    for k in new_keys - old_keys:
        row = {'变更类型': '[+] 新增', '位号': k}
        for col in new_df.columns: row[col] = str(new_df.loc[k, col])
        excel_data.append(row)

    # 2. 修改项与 DNP 拦截计算
    for key in old_keys & new_keys:
        old_row, new_row = old_df.loc[key], new_df.loc[key]
        row_data = {'变更类型': '[*] 修改', '位号': key}
        has_diff = False
        
        for col in list(set(old_df.columns) | set(new_df.columns)):
            in_old, in_new = col in old_df.columns, col in new_df.columns
            if in_old and in_new:
                val_old, val_new = old_row[col], new_row[col]
                if clean_value(val_old) != clean_value(val_new):
                    was_dnp, is_now_dnp = is_dnp(val_old), is_dnp(val_new)
                    if not was_dnp and is_now_dnp: row_data['变更类型'] = '[!] 警报: 变为空贴'
                    elif was_dnp and not is_now_dnp: row_data['变更类型'] = '[!] 警报: 恢复贴片'
                    row_data[col] = f"[{val_old}] -> [{val_new}]"
                    has_diff = True
            elif in_old and not in_new: row_data[col], has_diff = "[新版缺失此列]", True
            elif not in_old and in_new: row_data[col], has_diff = "[旧版缺失此列]", True
                
        if has_diff: excel_data.append(row_data)

    if not excel_data:
        messagebox.showinfo("核对完成", "完美匹配！两份 BOM 没有任何差异。")
        return

    result_df = pd.DataFrame(excel_data).fillna('')
    
    # ================= [防线升级：结果逆向重组与合并同类项 (Implode)] =================
    import re
    # 1. 提取所有非位号的列作为“聚类指纹”
    group_cols = [col for col in result_df.columns if col != '位号']
    
    # 2. 工业级自然排序算法 (解决 C10 排在 C2 前面的痛点)
    def natural_sort_key(s):
        return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(s))]
        
    def join_designators(series):
        """将散落的位号重新排序并用逗号压缩拼装"""
        return ', '.join(sorted(series.astype(str).tolist(), key=natural_sort_key))

    # 3. 核心大招：执行聚合压缩！
    # 把具有相同变更类型、相同Value、相同封装的器件强行拍扁成一行
    compressed_df = result_df.groupby(group_cols, as_index=False).agg({'位号': join_designators})
    # ===================================================================================

    # 重构表头顺序 (把变更类型和位号提到最前面)
    cols = compressed_df.columns.tolist()
    cols.insert(0, cols.pop(cols.index('变更类型')))
    cols.insert(1, cols.pop(cols.index('位号')))
    
    # 输出到最终 Excel
    compressed_df[cols].sort_values(by=['变更类型']).to_excel(output_filename, index=False)
    render_excel(output_filename, 'diff')

def process_lib_check(ad_path, lib_path, output_filename):
    """[模式 B]：AD BOM 与 公司 K3 库首次比对校验"""
    ad_df = load_bom_file(ad_path)
    lib_df = load_bom_file(lib_path) 

    if ad_df.empty or lib_df.empty: return
    if 'Designator' not in ad_df.columns or 'K3_Code' not in ad_df.columns:
        messagebox.showerror("识别失败", "AD BOM 中必须包含'位号'与'K3编码'相关列！")
        return
    if 'K3_Code' not in lib_df.columns:
        messagebox.showerror("识别失败", "公司 K3 库中未识别到'编码'列，请检查表头。")
        return

    # 1. 构建高速哈希字典 + Value/Footprint 倒排索引
    lib_dict = {}
    lib_code_lookup = {}
    candidate_index = {}
    value_index = {}
    candidate_code_rank = {}
    code_category = {}
    code_specs = {}
    for _, row in lib_df.iterrows():
        k3_code = str(row['K3_Code']).strip()
        if k3_code and k3_code != 'nan':
            specs = [str(row[c]).strip() for c in lib_df.columns if c != 'K3_Code' and str(row[c]).strip() not in ['', 'nan']]
            spec_text = " | ".join(specs)
            if k3_code not in lib_dict:
                candidate_code_rank[k3_code] = len(candidate_code_rank)
            lib_dict[k3_code] = spec_text
            code_category[k3_code] = get_lib_category(row, spec_text)
            code_specs[k3_code] = spec_text

            normalized_k3_code = normalize_k3_code_for_lookup(k3_code)
            if normalized_k3_code not in lib_code_lookup:
                lib_code_lookup[normalized_k3_code] = k3_code
            elif lib_code_lookup[normalized_k3_code] != k3_code:
                lib_code_lookup[normalized_k3_code] = None

            k3_value_text = str(row.get('Value', '')).strip()
            k3_foot_text = str(row.get('Footprint', '')).strip()
            if k3_value_text in ['', 'nan']:
                k3_value_text = spec_text
            if k3_foot_text in ['', 'nan']:
                k3_foot_text = spec_text

            foot_core = get_footprint_core(k3_foot_text)
            for value_key in build_value_search_keys(f"{k3_value_text} | {spec_text}"):
                add_candidate_index(candidate_index, value_key, foot_core, k3_code)
                add_candidate_index(value_index, value_key, '', k3_code)

    # 2. 链路状态指示灯：防御断路风险
    messagebox.showinfo("底层诊断", f"数据总线加载完毕！\n\nAD 待测物料数: {len(ad_df)} 项\nK3 标准库物料数: {len(lib_dict)} 项\n\n注：如果K3库数量极少，请确认导出的库是否包含了所有元件大类！")

    # 3. 三重校验 + Value/Footprint 候选编码反查
    excel_data = []
    candidate_cache = {}
    for _, row in ad_df.iterrows():
        k3_code = str(row.get('K3_Code', '')).strip()
        ad_val = str(row.get('Value', '')).strip()
        ad_foot = str(row.get('Footprint', '')).strip()
        ad_qty = str(row.get('Quantity', '')).strip() # <--- [新增] 提取数量数据
        
        row_data = {
            '位号 (Designator)': str(row.get('Designator', '')),
            '数量 (Qty)': ad_qty,  # <--- [新增] 将数量接入输出总线
            'AD K3编码': k3_code,
            'AD 原理图参数': f"Val: {ad_val} | Foot: {ad_foot}",
            'Value+Footprint匹配K3编码': '',
            '校验状态': '',
            '公司 K3 库标准参数': ''
        }
        # ... 后面的状态机和切片逻辑保持完全不变 ...

        # 状态机 A：缺件或彻底找不到
        if k3_code in ['', 'nan', 'NONE']:
            row_data['校验状态'] = '[!] 缺少编码'
        else:
            lib_k3_codes, missing_k3_codes = resolve_k3_codes(k3_code, lib_dict, lib_code_lookup)

        if row_data['校验状态']:
            pass
        elif missing_k3_codes or not lib_k3_codes:
            row_data['校验状态'] = '[x] 库中无此物料'
            
        # 状态机 B：编码存在 -> 启动深度交叉验证
        else:
            lib_specs_text = " | ".join(lib_dict[lib_code] for lib_code in lib_k3_codes)
            # ================= [核心终极修复：对称底噪抹平] =================
            # AD 的信号在底层被抹去了 Ω 和 ±，K3 库也必须同等抹去，否则会引发非对称断路！
            # ================================================================
            
            # ================= [防线升级 1] DNP 空贴噪声剥离与状态捕捉 =================
            is_row_dnp = False
            # 1. 检查是否存在专属的 DNP 列
            if 'DNP' in ad_df.columns and is_dnp(str(row.get('DNP', ''))):
                is_row_dnp = True
            # 2. 检查 Value 列是否夹杂了 DNP 字眼 (如 10K_DNP 或 0.1uF/NC)
            elif is_dnp(ad_val):
                is_row_dnp = True
                
            # DNP/NC/空贴 噪音会在 find_value_conflicts() 内部统一清洗
            # =========================================================================

            k3_results = []
            for lib_code in lib_k3_codes:
                match_val, match_foot, val_conflicts, tolerance_conflicts = check_k3_specs_match(
                    ad_val, ad_foot, row.get('Designator', ''), normalize_k3_specs(lib_dict[lib_code])
                )
                k3_results.append((lib_code, match_val, match_foot, val_conflicts, tolerance_conflicts))
            
            # --- 与门仲裁 ---
            # 如果捕捉到了 DNP 状态，准备好专属后缀标签
            dnp_tag = " (空贴/DNP)" if is_row_dnp else ""
            
            if all(match_val and match_foot for _, match_val, match_foot, _, _ in k3_results):
                row_data['校验状态'] = f'[√] 完美匹配{dnp_tag}'
            else:
                conflict_text = summarize_k3_mismatches(k3_results, ad_foot)
                row_data['校验状态'] = f'[!] K3编码不匹配({conflict_text}){dnp_tag}'
                
            row_data['公司 K3 库标准参数'] = lib_specs_text

        if row_data['校验状态'] and '[√]' not in row_data['校验状态']:
            candidate_key = (
                ad_val.upper(),
                ad_foot.upper(),
                get_designator_family(row.get('Designator', ''))
            )
            if candidate_key not in candidate_cache:
                matched_codes = find_value_footprint_k3_codes(
                    ad_val, ad_foot, row.get('Designator', ''),
                    candidate_index, value_index, candidate_code_rank, code_category, code_specs
                )
                candidate_cache[candidate_key] = '；'.join(matched_codes) or '未匹配'
            row_data['Value+Footprint匹配K3编码'] = candidate_cache[candidate_key]
            
        # ==================== [核心修复：焊接数据输出引脚] ====================
        excel_data.append(row_data) 
        # ======================================================================

    # 循环彻底结束后，再将收集到的数据转化为表格并排序
    result_df = pd.DataFrame(excel_data).fillna('')
    status_order = {'[!]': 0, '[x]': 1, '[√]': 2}
    result_df['_颜色分类排序'] = result_df['校验状态'].apply(
        lambda val: next((order for key, order in status_order.items() if key in str(val)), 9)
    )
    result_df.sort_values(
        by=['_颜色分类排序', 'Value+Footprint匹配K3编码', '校验状态', '位号 (Designator)'],
        inplace=True,
        kind='stable'
    )
    result_df.drop(columns=['_颜色分类排序'], inplace=True)
    result_df.to_excel(output_filename, index=False)
    render_excel(output_filename, 'check')

# ==================== Module 4: 渲染层 (莫兰迪护眼色彩驱动) ====================

def render_excel(filename, mode):
    wb = load_workbook(filename)
    ws = wb.active
    
    # 调色盘：启用低饱和度护眼色系 (莫兰迪色系)
    soft_red = PatternFill("solid", fgColor="FADBD8")    # 柔和粉红 (移除 / 找不到物料)
    soft_green = PatternFill("solid", fgColor="D5F5E3")  # 柔和薄荷绿 (新增 / 完美匹配)
    soft_yellow = PatternFill("solid", fgColor="FCF3CF") # 柔和奶黄 (数值修改)
    soft_orange = PatternFill("solid", fgColor="FDEBD0") # 柔和浅橙 (警报 / 参数冲突)
    
    # 柔和警报字体：使用砖红色替代刺眼的纯正红色
    alert_font = Font(bold=True, color="C0392B") 

    if mode == 'diff':
        fills = {'[-]': soft_red, '[+]': soft_green, '[!]': soft_orange, '[*]': soft_yellow}
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row=row, column=1).value)
            if '移除' in val: [setattr(c, 'fill', fills['[-]']) for c in ws[row]]
            elif '新增' in val: [setattr(c, 'fill', fills['[+]']) for c in ws[row]]
            elif '警报' in val: 
                for c in ws[row]: c.fill, c.font = fills['[!]'], alert_font
            elif '修改' in val:
                for c in ws[row]:
                    if c.value and '->' in str(c.value): c.fill = fills['[*]']

    elif mode == 'check':
        fills = {'[x]': soft_red, '[!]': soft_orange, '[√]': soft_green}
        headers = [str(cell.value) for cell in ws[1]]
        status_col = headers.index('校验状态') + 1 if '校验状态' in headers else 5
        for row in range(2, ws.max_row + 1):
            # [核心修正] 自动定位校验状态列，避免后续新增列导致色彩渲染错位
            val = str(ws.cell(row=row, column=status_col).value) 
            for key in fills:
                if key in val:
                    for c in ws[row]: 
                        c.fill = fills[key]
                        if key == '[!]': c.font = alert_font
                    break

    wb.save(filename)
    messagebox.showinfo("执行完毕", f"🎉 报表生成成功！\n已应用护眼色彩渲染，保存在:\n{filename}")

# ==================== Module 5: 人机交互界面 (自研机械导航引擎 + 动态说明书版) ====================

class GradientHeader(tk.Canvas):
    """自研 CPU 软件渲染器：通过逐行扫描绘制高拟真渐变背景"""
    def __init__(self, parent, color1, color2, **kwargs):
        super().__init__(parent, **kwargs)
        self._color1 = color1
        self._color2 = color2
        self.bind("<Configure>", self._draw_gradient)

    def _draw_gradient(self, event=None):
        self.delete("gradient")
        width, height = self.winfo_width(), self.winfo_height()
        if width <= 1 or height <= 1: return 
        
        r1, g1, b1 = self.winfo_rgb(self._color1)
        r2, g2, b2 = self.winfo_rgb(self._color2)
        r_ratio, g_ratio, b_ratio = (r2-r1)/width, (g2-g1)/width, (b2-b1)/width

        for i in range(width):
            nr, ng, nb = int(r1 + (r_ratio * i)), int(g1 + (g_ratio * i)), int(b1 + (b_ratio * i))
            color = "#%4.4x%4.4x%4.4x" % (nr, ng, nb)
            self.create_line(i, 0, i, height, tags=("gradient",), fill=color)
        self.tag_lower("gradient")

def show_instructions(root):
    """编译期固化说明书引擎：防篡改，仅开发者可通过重新打包修改"""
    help_win = tk.Toplevel(root)
    help_win.title("📖 Bom-Check使用说明")
    help_win.geometry("580x460")
    help_win.configure(bg="#EAECEE")
    
    # 设置模态窗口（置顶，且必须关闭后才能操作主界面）
    help_win.transient(root)
    help_win.grab_set()

    # ==================== [核心重构：防篡改的内置固化存储] ====================
    # 👨‍💻 开发者专属修改区：直接在这里修改文字，然后用 PyInstaller 重新打包即可！
    developer_instructions = """欢迎使用 Excite Bom-check HW V4.3!

【模式 A:硬件改版差异核对】
1. 作用：对比新旧两份 BOM,提取新增、移除、修改(包括变为空贴)的器件。
2. 注：两份表头必须包含“位号”"k3 no"列。

【模式 B:首次发板 K3 校验】
1. 作用：将 EDA 导出的 BOM 与公司 K3 库进行三维交叉比对(K3 No.+Value+Footprint)。
2. 高级特性：
   - 自动等效换算:100nF=0.1uF,1MR=1M 等。
   - DNP 状态剥离：自动从参数中剔除 DNP/NC/空贴 字符。(需要将DNP参数写在Value中)
   - 智能旁路路由：芯片(U)、电感(L)、二极管(D) 等物料,由于K3中不会存在封装信息,所以屏蔽封装检测,检测value与K3编码

=========================================
具体使用详情查看内部飞书文档
https://s17weazhe5w.feishu.cn/wiki/QDXNwjkjriEpdmkqfD0cvmALn2b
   
=========================================
⚠️ 内部管控工具，请勿外传。
如有增加新物料规则或功能需求，请联系工具开发者。"""
    # ==========================================================================

    # 渲染滚动文本视窗
    txt_frame = tk.Frame(help_win, bd=2, relief=tk.SUNKEN)
    txt_frame.pack(expand=True, fill='both', padx=20, pady=20)

    txt = tk.Text(txt_frame, font=("Microsoft YaHei", 10), bg="#FDFEFE", fg="#2C3E50", wrap="word", padx=10, pady=10)
    scrollbar = ttk.Scrollbar(txt_frame, command=txt.yview)
    txt.configure(yscrollcommand=scrollbar.set)
    
    scrollbar.pack(side='right', fill='y')
    txt.pack(side='left', expand=True, fill='both')
    
    # 将内置的固化文字插入到界面中
    txt.insert('1.0', developer_instructions)
    # [安全锁] 彻底锁定文本框，禁止用户在界面上敲击键盘篡改内容
    txt.config(state='disabled')

def run_app():
    root = tk.Tk()
    root.title("BOM-Check V4.3 ")
    root.geometry("760x540") # 稍微拉宽以容纳帮助按钮
    root.configure(bg="#EAECEE") 
    
    # ---------------- 1. 顶部炫酷渐变横幅 ----------------
    header_frame = tk.Frame(root, height=60)
    header_frame.pack(fill='x')
    header_frame.pack_propagate(False) 
    
    gradient = GradientHeader(header_frame, color1="#1A2980", color2="#26D0CE", highlightthickness=0)
    gradient.place(relwidth=1, relheight=1)
    tk.Label(header_frame, text="Excite BOM-Check HW V4.3", font=("Microsoft YaHei", 18, "bold"), 
             fg="white", bg="#1A2980").place(relx=0.5, rely=0.5, anchor="center")

    # ---------------- 2. 引入基础样式 ----------------
    style = ttk.Style()
    if 'clam' in style.theme_names(): style.theme_use('clam')
    font_main = ("Microsoft YaHei", 10)
    font_bold = ("Microsoft YaHei", 10, "bold")
    
    style.configure("TLabelframe", background="#FFFFFF", font=font_bold, foreground="#2E4053")
    style.configure("TLabelframe.Label", background="#FFFFFF")

    # ==================== [核心重构：带有帮助按钮的导航矩阵] ====================
    nav_frame = tk.Frame(root, bg="#EAECEE")
    nav_frame.pack(fill='x', pady=15, padx=20)
    
    # 建立帮助按钮 (靠右放置，小巧精致)
    btn_help = tk.Button(nav_frame, text="📖 使用说明", font=("Microsoft YaHei", 10, "bold"), 
                         fg="#2980B9", bg="#EAECEE", relief=tk.GROOVE, bd=2, cursor="hand2",
                         command=lambda: show_instructions(root))
    btn_help.pack(side='right', fill='y', padx=(10, 0))
    
    # 建立两个宏大的物理按键
    btn_a = tk.Button(nav_frame, text="🔄 模式 A：差异核对", font=("Microsoft YaHei", 12, "bold"), cursor="hand2")
    btn_b = tk.Button(nav_frame, text="🔍 模式 B：K3 校验", font=("Microsoft YaHei", 12, "bold"), cursor="hand2")
    
    btn_a.pack(side='left', expand=True, fill='x', padx=(0, 10), ipady=8)
    btn_b.pack(side='left', expand=True, fill='x', padx=(0, 0), ipady=8)
    
    # 建立主工作区容器
    work_area = tk.Frame(root, bg="#FFFFFF", bd=2, relief=tk.GROOVE)
    work_area.pack(fill='both', expand=True, padx=20, pady=(0, 20))
    
    tab_a = tk.Frame(work_area, bg="#FFFFFF")
    tab_b = tk.Frame(work_area, bg="#FFFFFF")

    # ======================== 选项卡 1：差异核对面板内容 ========================
    tk.Label(tab_a, text=" 应用场景：版本迭代时，精准捕获元器件的新增、移除、数值修改及空贴变化 ", 
             fg="#85929E", bg="#FFFFFF", font=("Microsoft YaHei", 9)).pack(pady=10)
    frame_a = ttk.LabelFrame(tab_a, text=" 导入配置 (Mode A) ")
    frame_a.pack(fill='x', padx=30, pady=5)
    
    v_old, v_new = tk.StringVar(), tk.StringVar()
    f1 = tk.Frame(frame_a, bg="#FFFFFF"); f1.pack(fill='x', padx=15, pady=8)
    tk.Button(f1, text="选择基准旧版", width=12, font=font_main, bg="#ECF0F1", relief=tk.GROOVE, bd=2,
              command=lambda: v_old.set(filedialog.askopenfilename())).pack(side='left')
    ttk.Entry(f1, textvariable=v_old, state='readonly').pack(side='left', fill='x', expand=True, padx=10)
    
    f2 = tk.Frame(frame_a, bg="#FFFFFF"); f2.pack(fill='x', padx=15, pady=8)
    tk.Button(f2, text="选择待测新版", width=12, font=font_main, bg="#ECF0F1", relief=tk.GROOVE, bd=2,
              command=lambda: v_new.set(filedialog.askopenfilename())).pack(side='left')
    ttk.Entry(f2, textvariable=v_new, state='readonly').pack(side='left', fill='x', expand=True, padx=10)
    
    tk.Button(tab_a, text="⚡ 开始执行差异核对", font=("Microsoft YaHei", 12, "bold"),
              bg="#2980B9", fg="white", activebackground="#1F618D", activeforeground="white", 
              relief=tk.RAISED, bd=5, cursor="hand2", 
              command=lambda: process_diff(v_old.get(), v_new.get(), os.path.join(get_base_dir(), "BOM_差异对比报告.xlsx"))).pack(pady=20, fill='x', padx=100)

    # ======================== 选项卡 2：K3 主数据校验面板内容 ========================
    tk.Label(tab_b, text=" 应用场景：新板打样前，三维交叉验证图纸器件是否在 K3 ERP 库中合法存在 ", 
             fg="#85929E", bg="#FFFFFF", font=("Microsoft YaHei", 9)).pack(pady=10)
    frame_b = ttk.LabelFrame(tab_b, text=" 导入配置 (Mode B) ")
    frame_b.pack(fill='x', padx=30, pady=5)

    v_ad, v_lib = tk.StringVar(), tk.StringVar()
    f3 = tk.Frame(frame_b, bg="#FFFFFF"); f3.pack(fill='x', padx=15, pady=8)
    tk.Button(f3, text="AD 导出 BOM", width=12, font=font_main, bg="#ECF0F1", relief=tk.GROOVE, bd=2,
              command=lambda: v_ad.set(filedialog.askopenfilename())).pack(side='left')
    ttk.Entry(f3, textvariable=v_ad, state='readonly').pack(side='left', fill='x', expand=True, padx=10)
    
    f4 = tk.Frame(frame_b, bg="#FFFFFF"); f4.pack(fill='x', padx=15, pady=8)
    tk.Button(f4, text="公司 K3 总库", width=12, font=font_main, bg="#ECF0F1", relief=tk.GROOVE, bd=2,
              command=lambda: v_lib.set(filedialog.askopenfilename())).pack(side='left')
    ttk.Entry(f4, textvariable=v_lib, state='readonly').pack(side='left', fill='x', expand=True, padx=10)

    tk.Button(tab_b, text="⚡ 开始三维交叉比对", font=("Microsoft YaHei", 12, "bold"),
              bg="#27AE60", fg="white", activebackground="#1D8348", activeforeground="white", 
              relief=tk.RAISED, bd=5, cursor="hand2", 
              command=lambda: process_lib_check(v_ad.get(), v_lib.get(), os.path.join(get_base_dir(), "K3物料校验报告.xlsx"))).pack(pady=20, fill='x', padx=100)

    # ==================== [导航引擎的底层路由逻辑] ====================
    def switch_to_a():
        btn_a.config(relief=tk.SUNKEN, bg="#2980B9", fg="white", activebackground="#2980B9")
        btn_b.config(relief=tk.RAISED, bg="#E5E7E9", fg="#A6ACAF", activebackground="#E5E7E9")
        tab_b.pack_forget() 
        tab_a.pack(fill='both', expand=True, padx=10, pady=10) 
        
    def switch_to_b():
        btn_b.config(relief=tk.SUNKEN, bg="#27AE60", fg="white", activebackground="#27AE60")
        btn_a.config(relief=tk.RAISED, bg="#E5E7E9", fg="#A6ACAF", activebackground="#E5E7E9")
        tab_a.pack_forget() 
        tab_b.pack(fill='both', expand=True, padx=10, pady=10) 

    btn_a.config(command=switch_to_a)
    btn_b.config(command=switch_to_b)
    switch_to_a() 
    
    root.mainloop()

if __name__ == "__main__":
    run_app()
